# api/export.py
import json
import os
import base64
from datetime import datetime
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import urllib.parse

# Your projects dictionary
PROJECTS = {
    "Component Library Project d3163287-84fc-432d-9c06-53709be1d545": "00eac9c4-b796-480b-bfe3-da6df473d190",
    "1710 - Solutions Projects": "f3506136-3dec-4345-937a-7a024dc9613c",
    "Project Controls Environment": "07e5c779-0698-4d47-b68e-17aabaae0d72",
    "1560 - Madinaty Open Air Center Phase 2": "3538977e-1c47-4b6b-971c-ccf7c9f188e9",
    "SQUARE BIM Project Template": "81c9d750-b7e0-45ff-b9d7-ad84b521fad2",
    "0000 - Tender Projects": "ce4cfb53-32f7-44ce-83ea-342397b2eda7",
    "0001 - Corporate Correspondence": "b0c426c6-324c-432f-ade9-5f0b832374a0",
    "2390 - Noor Bus Station": "dfe78031-a72d-47dd-ace4-17c22b49ff39",
    "SQUARE Noor Template": "d1d722bd-c74d-4fe4-8e67-a4090ecdd184",
    "Autodesk Test": "d00c1696-e10f-4410-906c-b34f365b39ea",
    "SQUARE Enhance Folders 2025": "ad5b8a5a-2e5f-4d55-b640-96a51d95f81e",
    "1590 - MADINATY V25 Villas": "d54ae1ab-5d0e-4514-951a-e4568c2edbdd",
    "2410 - Mena House Palace Renovation": "ff2e19a3-576c-46da-ae38-5f9e2ad411d3",
    "Training Project 2025": "b9e2030c-44f3-42d4-8bd7-18d78b3be468",
    "2220 - AL-Rehab Strip Mall": "f2aaf4e9-1c92-4edb-93aa-5c8b151d9b9f",
    "2810 - Noor Sport Club Phase 2": "08739574-00bc-4691-97ef-58d1f5fad5e6",
    "2140 - Madinaty Bus Station": "6e70243a-7781-4e46-9f81-0bf3393cd445",
    "2540 - SOUTH-MED T1-VILLAS": "38491092-4fa0-4e32-9f51-c80c5a95c4a5",
    "2150 - MADINATY Craft City Mall Commercial Buildings": "eece7526-0b7b-4b64-8003-e05f12ba9217",
    "1960 - MADINATY V24 Villas": "be44bce3-8a42-45cd-be0a-cbc0208ec27e",
    "1940 - MADINATY B12 District Center": "d49ae4ed-c04d-4458-a781-22c2eb364bc2",
    "1970 - MADINATY Commercial Strip": "c6150988-9021-429f-9683-b5858cb42135",
    "0000 - Reality Capture Projects": "6377d5ea-4438-4840-9b77-7bcf0511be3f",
    "Noor Sport Club": "7de33aa9-d0e9-4e9b-9db8-e408a5782e8e",
    "Sample Project - Seaport Civic Center": "1552c7dd-9709-4ed4-847d-915050067558",
    "SQUARE Prototype Project 2025": "c4639133-9399-4486-8b08-c54bcb1afed8",
    "1570 - MADINATY Open Air Mall - Building H": "b379a6f7-7dcd-48a3-a261-952da297bb9b",
    "1550 - Madinaty Water Park Mall": "dec4a7e3-adf5-4011-8f19-c4547059239f",
    "3010 - BANAN Residential Villas BV2.1": "2d5cef27-acde-4b86-a32f-9c399d85c19e",
    "Information Systems Workspace": "32d52d7d-b9fd-473d-8963-21ba503f854b",
    "2110 - MADINATY PRIVADO Central Park": "c4ee76c9-0ae4-4fa7-9a6e-3a351e297f4e",
    "2380 - Noor South Regional Commercial - Phase 1": "6bc8de17-3602-4a78-b7c1-6b2d4d0a6284",
    "2370 - Noor Data Center": "ab6daa2e-81a5-4fcb-a45f-e740f7821e31",
    "1580 - MADINATY V24 Service Center": "c4bae1fa-73b4-40e9-be93-0030166a5fce",
    "1610 - CELIA Villas V3": "708ebff3-10b0-4534-84a3-4b827b705b1c",
    "Project Control POC": "5c615ee2-2d53-4c01-a08c-da21ee0d119b",
    "2550 - SOUTH-MED Western Fence": "118fb7e1-27e7-4b19-9725-12ae89c27dda",
    "2320 - NOOR V4 Neighborhood Buildings": "3c2e9713-dcfd-4824-80df-5662be85673c",
    "2330 - Noor Commercial Center Phase 1": "baaf0b51-146b-48f0-9d19-095e3ac13b1a",
    "2530 - SOUTH-MED Site Office Building": "12e25623-df1f-44bd-9229-8717b96bc499",
    "2230 - North Eden AL-Rehab": "e5ec4264-ce5b-4f4c-9106-281eb5154ee5",
    "2350 - Noor School 1 & 2": "39842f1e-0ab9-4c82-9981-f4e53b5a0c30",
    "2210 - EDEN ALREHAB South of The Club": "2d4cfa38-2a37-4da5-8168-a99cc1a95f33",
    "2340 - Noor Sport Club Phase 1": "7bfda210-2347-450d-9437-e00fcf5692c5",
    "2360 - Noor Villas V4.4": "98042bf7-b28f-4755-9663-7f9a22274d0f",
    "2310 - NOOR V4.1 Villas Project": "df054120-a535-483b-8fd9-2cecb0ed5761",
    "2420 - Marriott MENA HOUSE Hotel Renovation": "934b18ec-fb59-4cf7-b99a-3171c317b6af",
    "2520 - SOUTH-MED V1-VILLAS": "0027f09a-0ea3-4327-a2ca-b0c789a064ef",
    "2620 - Marriott Zamalek Hotel Renovation": "175ea978-e543-44c8-9a0e-3b57e0ca5ac4",
    "3020 - BANAN Residential Villas BV4.1": "9e5399ed-4e25-45df-832e-1bd2905f4d43",
    "2190 - MADINATY PRIVADO Town Center": "64996a0a-b1b3-4b78-9bc7-61de21c97ae9",
    "1990 - MADINATY V35 Villas": "ba2013b6-0583-4ed5-a31a-ff9cc6a422b4",
    "1980 - AlThawra Building Project": "447fc161-6c5c-494d-af6a-7824aa83902e",
    "2920 - Aswan Old Cataract Hotel Renovation": "6bfb80b2-9d1a-494c-979b-ca10c9ad128e",
    "1820 - SHV Villas": "5c06f8af-5d5b-4071-a2f0-966f7b707856",
    "2930 - Aswan Steigenberger Hotel Renovation": "c84c8927-48bb-4dad-8d02-36e91c254d97",
    "1740 - Four Season Chalets Project": "498cdf3e-952f-4457-9897-8f8070240cd3",
    "2910 - Aswan Four Seasons Renovation": "6fe99714-c882-4cdd-a62f-3137d1a93823",
    "1410 - MADINATY Villas Repair": "c6d8b7a0-6343-4c5a-a59f-3ded649fc755",
    "2010 - CELIA The Village": "c128a9dc-b283-49e0-888c-2bfd9b59c280",
    "0000 - ACC Learning Hub": "69397ee6-7b2b-4cc5-8d1f-009993696a91"
}

# Get credentials from Vercel environment variables
CLIENT_ID = os.environ.get("AUTODESK_CLIENT_ID", "")
CLIENT_SECRET = os.environ.get("AUTODESK_CLIENT_SECRET", "")
BASE_URL = "https://developer.api.autodesk.com"


# Handler for Vercel serverless function
def handler(request):
    # Handle CORS preflight
    if request.method == "OPTIONS":
        return {
            "statusCode": 200,
            "headers": {
                "Access-Control-Allow-Origin": "*",
                "Access-Control-Allow-Methods": "POST, OPTIONS",
                "Access-Control-Allow-Headers": "Content-Type",
            },
            "body": ""
        }

    try:
        # Parse JSON data
        data = json.loads(request.body)

        refresh_token = data.get("refresh_token", "").strip()
        project_name = data.get("project_name", "").strip()

        if not refresh_token or not project_name:
            return {
                "statusCode": 400,
                "headers": {
                    "Content-Type": "application/json",
                    "Access-Control-Allow-Origin": "*"
                },
                "body": json.dumps({
                    "success": False,
                    "error": "Missing refresh_token or project_name"
                })
            }

        # Get project ID
        project_id = PROJECTS.get(project_name)
        if not project_id:
            return {
                "statusCode": 400,
                "headers": {
                    "Content-Type": "application/json",
                    "Access-Control-Allow-Origin": "*"
                },
                "body": json.dumps({
                    "success": False,
                    "error": f"Project not found: {project_name}"
                })
            }

        # 1. Get access token
        access_token = refresh_acc_token(refresh_token)

        # 2. Get planning issue types
        issue_types = get_issue_types(access_token, project_id)

        planning_id = None
        for item in issue_types.get("results", []):
            if item.get("title", "").lower() == "planning":
                planning_id = item.get("id")
                break

        if not planning_id:
            return {
                "statusCode": 400,
                "headers": {
                    "Content-Type": "application/json",
                    "Access-Control-Allow-Origin": "*"
                },
                "body": json.dumps({
                    "success": False,
                    "error": "No planning issues found in this project"
                })
            }

        # 3. Get issues (limit 50 for free tier)
        issues = get_issues(access_token, project_id, planning_id, limit=50)

        if not issues:
            return {
                "statusCode": 404,
                "headers": {
                    "Content-Type": "application/json",
                    "Access-Control-Allow-Origin": "*"
                },
                "body": json.dumps({
                    "success": False,
                    "error": "No issues to export"
                })
            }

        # 4. Create Excel file
        excel_data = create_excel_file(issues, project_name)
        excel_base64 = base64.b64encode(excel_data).decode("utf-8")

        # Return success
        return {
            "statusCode": 200,
            "headers": {
                "Content-Type": "application/json",
                "Access-Control-Allow-Origin": "*"
            },
            "body": json.dumps({
                "success": True,
                "filename": f'{project_name.replace(" ", "_")}_issues_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx',
                "excel_file": excel_base64,
                "issue_count": len(issues),
                "project_name": project_name
            })
        }

    except Exception as e:
        return {
            "statusCode": 500,
            "headers": {
                "Content-Type": "application/json",
                "Access-Control-Allow-Origin": "*"
            },
            "body": json.dumps({
                "success": False,
                "error": str(e)
            })
        }


# Helper functions
def refresh_acc_token(refresh_token):
    """Get new access token using refresh token"""
    url = f"{BASE_URL}/authentication/v2/token"
    payload = {
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "grant_type": "refresh_token",
        "refresh_token": refresh_token
    }

    response = requests.post(url, data=payload, timeout=30)
    response.raise_for_status()
    return response.json()["access_token"]


def get_issue_types(access_token, project_id):
    """Get issue types for project"""
    url = f"{BASE_URL}/construction/issues/v1/projects/{project_id}/issue-types?include=subtypes"
    headers = {"Authorization": f"Bearer {access_token}"}

    response = requests.get(url, headers=headers, timeout=30)
    response.raise_for_status()
    return response.json()


def get_issues(access_token, project_id, planning_id, limit=50):
    """Get planning issues"""
    url = f"{BASE_URL}/construction/issues/v1/projects/{project_id}/issues"
    headers = {"Authorization": f"Bearer {access_token}"}
    params = {
        "limit": limit,
        "filter[issueTypeId]": planning_id
    }

    response = requests.get(url, headers=headers, params=params, timeout=30)
    response.raise_for_status()
    return response.json().get("results", [])


def create_excel_file(issues, project_name):
    """Create Excel file from issues"""
    wb = Workbook()
    ws = wb.active

    # Title
    ws["A1"] = f"{project_name} - ACC Issue Log"
    ws["A1"].font = Font(bold=True, size=14)

    # Date
    ws["A2"] = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    # Headers
    headers = ["No.", "Title", "Description", "Status", "Created", "Updated"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Add issues
    for i, issue in enumerate(issues, start=1):
        row = 4 + i

        ws.cell(row=row, column=1, value=i)  # No.
        ws.cell(row=row, column=2, value=issue.get("title", ""))  # Title
        ws.cell(row=row, column=3, value=issue.get("description", ""))  # Description
        ws.cell(row=row, column=4, value=issue.get("status", ""))  # Status

        # Format dates
        created = issue.get("createdAt", "")
        if created:
            ws.cell(row=row, column=5, value=created[:10])  # Created date

        updated = issue.get("updatedAt", "")
        if updated:
            ws.cell(row=row, column=6, value=updated[:10])  # Updated date

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to bytes
    import io
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()